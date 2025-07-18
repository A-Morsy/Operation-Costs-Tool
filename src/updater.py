import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from name_mappings import ALIAS_TO_CANONICAL, normalize_name, SKIP_NAMES
import os

class EmployeeCostMapper:
    def __init__(self, target_path, processed_path, target_sheet='Project', skip_names=None):
        self.target_path = target_path
        self.processed_path = processed_path
        self.target_sheet = target_sheet
        self.normalize_name = normalize_name
        self.processed_df = pd.read_excel(self.processed_path)
        # Load workbook for updating (formulas)
        self.wb = load_workbook(self.target_path,data_only=True)
        self.ws = self.wb[self.target_sheet]
        # Load workbook for values (data_only=True)
        self.wb_values = load_workbook(self.target_path, data_only=True)
        self.ws_values = self.wb_values[self.target_sheet]
        self.month_col_map = self._map_month_columns()
        self.operation_names = set(self.processed_df['Operation'].unique())
        self.processed_df['Name_norm'] = self.processed_df['Name'].apply(self.normalize_name)
        self.processed_df['Operation_norm'] = self.processed_df['Operation'].str.strip()
        self.skip_names = set(n.strip().lower() for n in (skip_names or []))
        self.mapping = []
        

    def _map_month_columns(self):
        """Map month names to their column indices."""
        month_col_map = {}
        for col in range(6, self.ws.max_column + 1):  # Month columns start at F (no.6)
            cell_value = self.ws.cell(row=1, column=col).value
            if cell_value:
                # Normalize to 'YYYY-MM' string
                if isinstance(cell_value, str) and cell_value[:4].isdigit():
                    period = cell_value[:7]
                elif hasattr(cell_value, 'strftime'):
                    period = cell_value.strftime('%Y-%m')
                else:
                    continue
                month_col_map[period] = col
        return month_col_map

    def _is_operation(self, value):
        return value and str(value).strip() in self.operation_names

    def _should_skip(self, value):
        return value and str(value).strip().lower() in self.skip_names

    def map_employees(self):
        self.cell_audit_log = []
        current_operation = None
        for row in range(2, self.ws.max_row + 1):
            cell_value = self.ws.cell(row=row, column=5).value  # Column E

            # Stop processing if "Accumulated Total" is reached
            if cell_value and str(cell_value).strip().lower() == "accumulated total":
                print(f"Reached 'Accumulated Total' at row {row}. Stopping iteration.")
                break

            # Detect operation section
            if self._is_operation(cell_value):
                current_operation = str(cell_value).strip()
                continue

            if not current_operation:
                continue

            emp_name = self.normalize_name(cell_value)
            if not emp_name or emp_name.strip() == '' or self._is_operation(emp_name) or self._should_skip(emp_name):
                continue

            for period, col in self.month_col_map.items():
                cell = self.ws.cell(row=row, column=col)
                cell_address = f"{get_column_letter(col)}{row}"
                # Use the value from the value-only workbook for Fees_before
                before_val = self.ws_values.cell(row=row, column=col).value
                # Set Fees_before to -1 if missing
                if before_val is None or (isinstance(before_val, float) and pd.isna(before_val)):
                    before_val = -1

                # Store BEFORE update. After update will be filled in next step.
                self.cell_audit_log.append({
                    'Operation': current_operation,
                    'Employee': emp_name.strip(),
                    'Month': period,
                    'Cell': cell_address,
                    'Fees_before': before_val,
                    'Fees_after': None  # Filled after updating
                })
        # Done mapping! self.cell_audit_log = full before-state

    def update_costs(self):
        orange_font = Font(color="FFA500")
        update_count = 0
        for entry in self.cell_audit_log:
            # Find the correct fee in the processed DataFrame
            mask = (
                (self.processed_df['Operation_norm'] == entry['Operation']) &
                (self.processed_df['Name_norm'] == entry['Employee']) &
                (self.processed_df['Month'] == entry['Month'])
            )
            match = self.processed_df[mask]
            if not match.empty:
                fee = match['Fees'].values[0]
                cell = self.ws[entry['Cell']]
                prev_val = cell.value
                cell.value = fee
                cell.font = orange_font
                update_count += 1
                # Record the after value
                entry['Fees_after'] = fee if fee is not None and not (isinstance(fee, float) and pd.isna(fee)) else -1
            else:
                # No match in processed data: Fees_after should be -1 to indicate missing
                entry['Fees_after'] = -1
        print(f"Updated {update_count} cells with new costs (highlighted in orange).")
        self.wb.save(self.target_path.replace('.xlsx', '_updated.xlsx'))

    def save_audit_chunks(self, output_dir="../chunks/target_chunks"):
        """Save audit log as chunks (by operation/project) with before/after fees."""
        all_df = pd.DataFrame(self.cell_audit_log)
        os.makedirs(output_dir, exist_ok=True)
        for op, chunk in all_df.groupby('Operation'):
            # Remove rows with NaN in Fees_before or Fees_after before saving
            chunk_clean = chunk.dropna(subset=['Fees_before', 'Fees_after'])
            safe_op = op.replace(' ', '_').replace('/', '_')
            path = os.path.join(output_dir, f"[audit]_{safe_op}.xlsx")
            chunk_clean.to_excel(path, index=False)
            print(f"Saved audit chunk: {path}")


if __name__ == "__main__":
    import sys
    # Get absolute paths relative to this script's directory
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    TARGET_FILE = os.path.abspath(os.path.join(BASE_DIR, "../docs/target_sheet.xlsx"))
    PROCESSED_FILE = os.path.abspath(os.path.join(BASE_DIR, "../docs/processed_timesheet.xlsx"))
    SHEET_NAME = "Project"

    # Print working directory and resolved paths for debug purposes
    print(f"Current working directory: {os.getcwd()}")
    print(f"Resolved TARGET_FILE: {TARGET_FILE}")
    print(f"Resolved PROCESSED_FILE: {PROCESSED_FILE}")

    mapper = EmployeeCostMapper(TARGET_FILE, PROCESSED_FILE, SHEET_NAME, SKIP_NAMES)
    mapper.map_employees()
    mapper.update_costs()
    mapper.save_audit_chunks()